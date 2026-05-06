import argparse
import csv
import logging
import os
from datetime import datetime
from io import BytesIO

import requests
from openpyxl import load_workbook

from scrape_itaipu import REQUEST_HEADERS

logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

CATEGORIA_SUFFIX = " BASICO CAT."
DATE_FORMATS = ["%d/%m/%Y", "%d/%m/%y", "%m/%d/%y", "%m/%d/%Y"]

CSV_HEADER = [
    "cedula", "nombre", "fecha_de_admision", "categoria", "salario", "sede",
    "antiguedad_eby", "antiguedad", "titulo", "zo", "desa", "ayuda",
    "he", "bonificaciones",
]


def fetch_xlsx_sheet(url):
    logging.info(f"Fetching xlsx {url=}")
    r = requests.get(url, headers=REQUEST_HEADERS)
    r.raise_for_status()
    wb = load_workbook(BytesIO(r.content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise RuntimeError(f"No active worksheet in {url}")
    return ws


def iter_dict_rows(ws, header_row=0):
    rows = ws.iter_rows(values_only=True)
    for _ in range(header_row):
        next(rows)
    headers = next(rows)
    for row in rows:
        yield dict(zip(headers, row))


def fetch_salary_scale(url):
    ws = fetch_xlsx_sheet(url)
    scale: dict[str, int] = {}
    for r in iter_dict_rows(ws):
        if r["categoria"] is None or r["basico"] is None:
            continue
        cat = str(r["categoria"]).strip()
        if cat.endswith(CATEGORIA_SUFFIX):
            cat = cat[: -len(CATEGORIA_SUFFIX)].strip()
        scale[cat] = int(str(r["basico"]))
    logging.info(f"Loaded {len(scale)} salary-scale entries")
    return scale


def normalize_date(raw):
    raw = str(raw).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    raise ValueError(f"Could not parse date: {raw}")


def resolve_basico(raw, salary_scale):
    raw_str = str(raw).strip()
    try:
        return "", int(raw_str)
    except ValueError:
        pass
    if raw_str not in salary_scale:
        raise KeyError(f"Categoria {raw_str!r} not found in salary scale")
    return raw_str, salary_scale[raw_str]


def _norm(v):
    return str(v).strip() if v is not None else ""


def fetch_nomina_base(url):
    """WEB2 — canonical nomina. Returns dict keyed by cedula and the cedula order."""
    ws = fetch_xlsx_sheet(url)
    base: dict[str, dict] = {}
    order: list[str] = []
    for r in iter_dict_rows(ws):
        if _norm(r["cedula"]) == "":
            logging.info(f"Skipping WEB2 row with empty cedula: {r}")
            continue
        cedula = _norm(r["cedula"]).replace(".", "")
        base[cedula] = {
            "nombre": _norm(r["nombre"]),
            "basico": _norm(r["basico"]),
            "ingreso": _norm(r["ingreso"]),
            "sede": _norm(r["sede"]),
        }
        order.append(cedula)
    return base, order


def fetch_bonificaciones(url):
    """WEB3 — bonus data plus a copy of the base columns (used for discrepancy checks)."""
    ws = fetch_xlsx_sheet(url)
    out: dict[str, dict] = {}
    # row 0: "Mes: <month>" snapshot marker; row 1: blank; row 2: headers
    for r in iter_dict_rows(ws, header_row=2):
        if _norm(r["cedula"]) == "":
            logging.info(f"Skipping WEB3 row with empty cedula: {r}")
            continue
        cedula = _norm(r["cedula"]).replace(".", "")
        out[cedula] = {
            "nombre": _norm(r["nombre"]),
            "basico": _norm(r["basico"]),
            "ingreso": _norm(r["ingreso"]),
            "sede": _norm(r["sede"]),
            "antiguedad_eby": int(str(r["antiguedad"] or 0)),
            "antiguedad": int(str(r["antigueda2"] or 0)),
            "titulo": int(str(r["titulo"] or 0)),
            "zo": int(str(r["zo"] or 0)),
            "desa": int(str(r["desa"] or 0)),
            "ayuda": int(str(r["ayuda"] or 0)),
            "he": int(str(r["he"] or 0)),
            "bonificaciones": int(str(r["bonificaciones"] or 0)),
        }
    return out


SHARED_FIELDS = ("nombre", "basico", "ingreso", "sede")
ZERO_BONUS = {k: 0 for k in (
    "antiguedad_eby", "antiguedad", "titulo", "zo", "desa", "ayuda",
    "he", "bonificaciones",
)}


def join_nominas(base, order, bonus, salary_scale):
    only_in_web3 = sorted(set(bonus) - set(base))
    if only_in_web3:
        logging.warning(
            f"{len(only_in_web3)} cedula(s) in WEB3 but not WEB2 (skipping): {only_in_web3[:10]}"
        )

    discrepancies = 0
    missing_in_web3 = 0
    for cedula in order:
        b = base[cedula]
        bo = bonus.get(cedula)
        if bo is None:
            missing_in_web3 += 1
            logging.warning(f"cedula {cedula} in WEB2 but not WEB3; emitting zeros for bonus columns")
            bo = ZERO_BONUS
        else:
            for f in SHARED_FIELDS:
                if b[f] != bo[f]:
                    discrepancies += 1
                    logging.warning(
                        f"cedula {cedula} {f} differs: WEB2={b[f]!r} WEB3={bo[f]!r} — using WEB2"
                    )
        categoria, salario = resolve_basico(b["basico"], salary_scale)
        yield [
            cedula,
            b["nombre"],
            normalize_date(b["ingreso"]),
            categoria,
            str(salario),
            b["sede"],
            str(bo["antiguedad_eby"]),
            str(bo["antiguedad"]),
            str(bo["titulo"]),
            str(bo["zo"]),
            str(bo["desa"]),
            str(bo["ayuda"]),
            str(bo["he"]),
            str(bo["bonificaciones"]),
        ]
    logging.info(
        f"Join complete: {len(order)} WEB2 rows, "
        f"{missing_in_web3} missing in WEB3, "
        f"{len(only_in_web3)} extra in WEB3, "
        f"{discrepancies} shared-field discrepancies"
    )


def main():
    parser = argparse.ArgumentParser(description="Scrape EBY/yacyreta nomina + bonificaciones to CSV")
    parser.add_argument(
        "--web1-url", default="https://www.eby.gov.py/sueldos/nominas/assets/WEB1.xlsx",
        help="WEB1 (escala salarial) URL",
    )
    parser.add_argument(
        "--web2-url", default="https://www.eby.gov.py/sueldos/nominas/assets/WEB2.xlsx",
        help="WEB2 (canonical nomina) URL",
    )
    parser.add_argument(
        "--web3-url", default="https://www.eby.gov.py/sueldos/nominas/assets/WEB3.xlsx",
        help="WEB3 (bonificaciones) URL",
    )
    parser.add_argument(
        "--output", default=None,
        help="CSV output path (default: data/yacyreta/nomina_yacyreta_<TODAY>.csv)",
    )
    args = parser.parse_args()

    salary_scale = fetch_salary_scale(args.web1_url)
    base, order = fetch_nomina_base(args.web2_url)
    bonus = fetch_bonificaciones(args.web3_url)
    rows = list(join_nominas(base, order, bonus, salary_scale))

    filename = args.output or os.path.join(
        "data/yacyreta",
        f"nomina_yacyreta_{datetime.now().strftime('%Y-%m-%d')}.csv",
    )
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    logging.info(f"Writing {len(rows)} rows to {filename}")
    with open(filename, "w") as f:
        writer = csv.writer(f)
        writer.writerow(CSV_HEADER)
        for r in rows:
            writer.writerow(r)


if __name__ == "__main__":
    main()
